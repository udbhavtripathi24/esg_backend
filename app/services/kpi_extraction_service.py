"""Layer 1: extraction service — reads an approved dataset version's data
file(s) and writes real, raw, as-reported KpiValue rows.

CRITICAL SAFETY DESIGN: this function parses and validates EVERYTHING into
memory first, and only touches the database session ONCE, via a single
add_all() at the very end, after every row has been validated. This is
deliberate: the calling worker (app/workers/__init__.py's process_one)
does NOT call session.rollback() on handler failure — it re-uses the same
session to update the job's own status. If this function partially
inserted rows via the session and then raised, the outer code's job-status
commit could commit that partial, inconsistent state alongside a "failed"
job record. Parsing fully into memory first, and only adding at the end
right before returning, makes a hard failure here leave the session
completely untouched — safe under the existing worker's actual behavior,
not just in the ideal case.

BOUNDARY, enforced by design, not just documentation: this function never
converts a unit, never applies an emission factor, never computes a
score. It extracts exactly what a human reported, verbatim, with full
traceability back to the source row. See app/models/kpi.py for the full
architectural rationale.
"""
import logging
from datetime import datetime
from openpyxl import load_workbook
from sqlmodel import Session, select

from app.models.dataset import Dataset, DatasetVersion, DatasetFile
from app.models.upload_type import UploadType
from app.models.kpi import KpiDefinition, KpiValue
from app.models.master_data import Site
from app.storage.factory import get_storage

log = logging.getLogger("kpi_extraction")


class ExtractionError(Exception):
    """Raised for a hard failure — file unreadable, dataset/version missing,
    etc. NOT raised for ordinary per-row data-quality issues (those are
    skipped with a warning, not a hard failure)."""


# Column header keyword matching. Deliberately lenient/keyword-based rather
# than exact-string, since domainGuidance's real column names include
# parenthetical examples (e.g. "Energy type (electricity, gas, diesel)")
# that a real uploader's header cell may or may not reproduce verbatim.
# This is provisional v1 engineering, consistent with domainGuidance's own
# approved-as-provisional status — not a claim of final methodology.
_SITE_KEYWORDS = ["site"]
_PERIOD_KEYWORDS = ["period"]

_DOMAIN_COLUMN_MAP = {
    "energy_data": {
        "value_cols": [("consumption", "energy.consumption", "energy_type", ["energy type", "energy_type"])],
    },
    "water_data": {
        "value_cols": [
            ("withdrawn", "water.withdrawal", "water_source_type", ["source type", "source_type"]),
            ("recycled", "water.recycled", "water_source_type", ["source type", "source_type"]),
        ],
    },
    "emissions_data": {
        "value_cols": [("activity data", "emissions.activity_data", "emission_scope", ["scope"])],
    },
    "waste_data": {
        "value_cols": [("quantity", "waste.generated", "waste_type", ["waste type", "waste_type"])],
    },
}
# Waste additionally carries a second attribute (disposal method) that
# qualifies the same value rather than being its own measurement.
_WASTE_DISPOSAL_KEYWORDS = ["disposal"]
_UNIT_KEYWORDS = ["unit"]


def _find_col(header_row, keywords):
    """Return the 0-based column index whose header contains ANY of the
    given keywords (case-insensitive substring match), or None."""
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        text = str(cell).strip().lower()
        if any(kw in text for kw in keywords):
            return idx
    return None


def extract_kpi_values_for_version(session: Session, dataset_version_id: int, job_id: int | None = None) -> dict:
    """Entry point, called from the recalculate_kpi_values worker job.

    Returns a summary dict: {"rows_written": N, "rows_skipped": N, "reason": ...}
    Raises ExtractionError only for hard failures (missing records, unreadable file).
    """
    version = session.get(DatasetVersion, dataset_version_id)
    if not version:
        raise ExtractionError(f"DatasetVersion {dataset_version_id} not found")

    # Idempotency: if this version was already extracted, don't redo it.
    # Checked BEFORE any parsing work, per the safety design above.
    existing = session.exec(
        select(KpiValue).where(KpiValue.dataset_version_id == dataset_version_id).limit(1)
    ).first()
    if existing:
        log.info(f"kpi_values already exist for dataset_version_id={dataset_version_id}, skipping")
        return {"rows_written": 0, "rows_skipped": 0, "reason": "already_extracted"}

    dataset = session.get(Dataset, version.dataset_id)
    if not dataset:
        raise ExtractionError(f"Dataset {version.dataset_id} not found")

    upload_type = session.get(UploadType, dataset.upload_type_id)
    if not upload_type:
        raise ExtractionError(f"UploadType {dataset.upload_type_id} not found")

    domain_map = _DOMAIN_COLUMN_MAP.get(upload_type.code)
    if not domain_map:
        # Genuinely unsupported domain (e.g. general_evidence, which is
        # never parsed for numbers per the existing, already-approved
        # decision). Not an error — just nothing to extract.
        return {"rows_written": 0, "rows_skipped": 0, "reason": "no_kpi_mapping_for_upload_type"}

    data_files = session.exec(
        select(DatasetFile).where(
            DatasetFile.dataset_version_id == dataset_version_id,
            DatasetFile.role == "data",
        )
    ).all()
    if not data_files:
        return {"rows_written": 0, "rows_skipped": 0, "reason": "no_data_file"}

    # Active KPI definitions for this upload type — used to validate every
    # kpi_code before writing (application-level referential integrity,
    # see app/models/kpi.py's note on why kpi_code isn't a hard FK), and
    # to stamp the exact active version onto every written row (hardening:
    # see KpiValue.kpi_definition_version's docstring for why this matters).
    definition_rows = session.exec(
        select(KpiDefinition.code, KpiDefinition.version).where(
            KpiDefinition.upload_type_id == upload_type.id,
            KpiDefinition.is_active == True,  # noqa: E712
        )
    ).all()
    valid_codes = {code: version for code, version in definition_rows}

    # Site lookup, scoped to this dataset's company — never fabricated;
    # a row whose site text doesn't match any real Site simply gets
    # site_id=None rather than a guessed value.
    sites_by_name = {
        s.name.strip().lower(): s.id
        for s in session.exec(select(Site).where(Site.company_id == dataset.company_id)).all()
    }
    sites_by_code = {
        s.code.strip().lower(): s.id
        for s in session.exec(select(Site).where(Site.company_id == dataset.company_id)).all()
    }

    storage = get_storage()
    to_insert: list[KpiValue] = []
    rows_skipped = 0

    for f in data_files:
        try:
            file_stream = storage.get(f.storage_key)
        except Exception as e:
            raise ExtractionError(f"Could not read file {f.storage_key}: {e}")

        try:
            wb = load_workbook(file_stream, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            raise ExtractionError(f"Could not parse workbook {f.original_filename}: {e}")

        if not rows:
            continue
        header = rows[0]
        site_col = _find_col(header, _SITE_KEYWORDS)
        unit_col = _find_col(header, _UNIT_KEYWORDS)
        disposal_col = _find_col(header, _WASTE_DISPOSAL_KEYWORDS) if upload_type.code == "waste_data" else None

        for row_idx, row in enumerate(rows[1:], start=2):  # 1-based, skip header
            if row is None or all(c is None for c in row):
                continue  # blank row, not an error

            site_text = str(row[site_col]).strip().lower() if site_col is not None and row[site_col] is not None else None
            site_id = sites_by_name.get(site_text) or sites_by_code.get(site_text) if site_text else None
            unit_val = str(row[unit_col]).strip() if unit_col is not None and row[unit_col] is not None else "unspecified"

            for value_keyword, kpi_code, attr_key, attr_keywords in domain_map["value_cols"]:
                if kpi_code not in valid_codes:
                    log.warning(f"kpi_code '{kpi_code}' has no active KpiDefinition — skipping")
                    rows_skipped += 1
                    continue

                value_col = _find_col(header, [value_keyword])
                if value_col is None:
                    rows_skipped += 1
                    continue

                raw_value = row[value_col]
                try:
                    numeric_value = float(raw_value)
                except (TypeError, ValueError):
                    # Invalid/missing data — skip this specific metric for
                    # this row, not a hard failure of the whole extraction.
                    rows_skipped += 1
                    continue

                attr_col = _find_col(header, attr_keywords)
                attributes = {}
                if attr_col is not None and row[attr_col] is not None:
                    attributes[attr_key] = str(row[attr_col]).strip()
                if disposal_col is not None and row[disposal_col] is not None:
                    attributes["disposal_method"] = str(row[disposal_col]).strip()

                to_insert.append(KpiValue(
                    dataset_id=dataset.id,
                    dataset_version_id=version.id,
                    source_file_id=f.id,
                    source_row_number=row_idx,
                    extraction_job_id=job_id,
                    company_id=dataset.company_id,
                    site_id=site_id,
                    kpi_code=kpi_code,
                    kpi_definition_version=valid_codes[kpi_code],
                    value=numeric_value,
                    unit=unit_val,
                    attributes=attributes,
                    reporting_period_start=dataset.reporting_period_start,
                    reporting_period_end=dataset.reporting_period_end,
                ))

    # Single, atomic write — see module docstring for why this happens
    # only once, at the very end. Delegated to _write_kpi_values (below)
    # so its DB-level conflict-handling can be tested independently of
    # this function's own earlier, application-level pre-check.
    return _write_kpi_values(session, to_insert, rows_skipped)


def _write_kpi_values(session: Session, to_insert: list[KpiValue], rows_skipped: int) -> dict:
    """Isolated so the DB-level conflict-handling can be tested directly,
    independent of the higher-level 'already extracted' pre-check in
    extract_kpi_values_for_version (which is a separate, earlier guard —
    see that function's own docstring note on why both layers exist).

    HARDENING (post-launch review): the caller's early "already extracted"
    check is an application-level pre-check, not a database-level
    guarantee — it cannot prevent a genuine race if two workers somehow
    ran extraction for the same version at the same moment (both could
    pass that check before either commits). The REAL safety net is the
    unique constraint on kpi_values itself (uq_kpi_value_source_row_metric).
    If that constraint is ever violated here, it means a concurrent run
    already wrote these exact rows between the pre-check and this commit
    — an explicit rollback and a clean, understood "already extracted
    concurrently" outcome is correct, not a crash. Critically, the
    rollback happens HERE, before returning — matching this module's own
    documented safety design of never leaving the session dirty for the
    outer worker's job-status commit to inherit.
    """
    if to_insert:
        from sqlalchemy.exc import IntegrityError
        session.add_all(to_insert)
        try:
            session.commit()
        except IntegrityError:
            session.rollback()
            log.warning(
                "kpi_values unique constraint hit — another process already "
                "extracted this concurrently. Treating as already-extracted."
            )
            return {"rows_written": 0, "rows_skipped": 0, "reason": "already_extracted_concurrently"}
    return {"rows_written": len(to_insert), "rows_skipped": rows_skipped, "reason": "extracted"}
