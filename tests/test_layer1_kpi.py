"""Layer 1 tests — KPI extraction, provenance, tenancy, idempotency,
invalid-data handling, and integration with the existing review lifecycle.

Follows the exact same pattern already established in test_worker.py:
handlers are called DIRECTLY with the test's own session fixture, since
process_one() opens its own separate Session(engine) which is incompatible
with the test suite's in-memory SQLite fixture (see test_worker.py's own
docstring for why).
"""
import io
from datetime import date
from sqlmodel import select
import pytest

from tests.conftest_helpers import bootstrap, make_company, make_user, auth
from scripts_seed_upload_types import seed_upload_types
from scripts_seed_kpi_definitions import seed_kpi_definitions
from app.models.dataset import Dataset, DatasetVersion, DatasetFile
from app.models.processing_job import ProcessingJob
from app.models.upload_type import UploadType
from app.models.kpi import KpiDefinition, KpiValue
from app.models.master_data import Site


# ---------- Realistic fixture builders, matching real domainGuidance columns ----------

def _energy_xlsx(rows) -> bytes:
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(["Site name or ID", "Energy type (electricity, gas, diesel)",
               "Consumption value", "Unit (kWh, GJ, L)", "Reporting period"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _water_xlsx(rows) -> bytes:
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(["Site name or ID", "Water source type", "Volume withdrawn",
               "Volume recycled", "Unit (ML, m3)", "Reporting period"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _waste_xlsx(rows) -> bytes:
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(["Site name or ID", "Waste type", "Quantity generated",
               "Disposal method", "Unit (t, kg)", "Reporting period"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _emissions_xlsx(rows) -> bytes:
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(["Site name or ID", "Emission scope", "Activity data",
               "Emission factor used", "Unit (tCO2e)", "Reporting period"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _set_storage(tmp_path):
    from app.storage.factory import get_storage as _gs
    _gs.cache_clear()
    from app.core.config import settings
    settings.STORAGE_LOCAL_ROOT = str(tmp_path)
    settings.STORAGE_BACKEND = "local"


def _seed(session):
    org = bootstrap(session)
    seed_upload_types(session)
    seed_kpi_definitions(session)
    return org


def _make_site(session, company, code, name):
    s = Site(company_id=company.id, code=code, name=name)
    session.add(s); session.commit(); session.refresh(s)
    return s


def _submitted_version(client, session, uploader, co, upload_type_code, file_bytes, filename="d.xlsx"):
    r = client.post("/api/v1/datasets",
                    json={"company_id": co.id, "upload_type_code": upload_type_code,
                          "reporting_period_start": "2026-04-01",
                          "reporting_period_end": "2026-06-30"},
                    headers=auth(uploader))
    assert r.status_code == 201, r.text
    ds_pid = r.json()["public_id"]
    v_pid = client.get(f"/api/v1/datasets/{ds_pid}/versions", headers=auth(uploader)).json()[0]["public_id"]
    r = client.post(f"/api/v1/datasets/{ds_pid}/versions/{v_pid}/files",
                    files={"file": (filename, file_bytes,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                    data={"role": "data"}, headers=auth(uploader))
    assert r.status_code == 201, r.text
    r = client.post(f"/api/v1/datasets/{ds_pid}/versions/{v_pid}/submit", headers=auth(uploader))
    assert r.status_code == 200, r.text
    return ds_pid, v_pid


def _approve(client, session, uploader, reviewer, ds_pid, v_pid):
    r = client.post(f"/api/v1/datasets/{ds_pid}/versions/{v_pid}/reviews",
                    json={"reviewer_user_id": reviewer.id}, headers=auth(uploader))
    assert r.status_code == 201, r.text
    rv_pid = r.json()["public_id"]
    r = client.post(f"/api/v1/datasets/{ds_pid}/versions/{v_pid}/reviews/{rv_pid}/decide",
                    json={"decision": "approved", "note": "Looks good."}, headers=auth(reviewer))
    assert r.status_code == 200, r.text


def _run_kpi_job(session, dataset_version_public_id):
    """Directly invoke the handler with the test's own session -- matches
    the exact established pattern in test_worker.py."""
    v = session.exec(select(DatasetVersion).where(
        DatasetVersion.public_id == dataset_version_public_id)).first()
    session.refresh(v)
    job = session.exec(select(ProcessingJob).where(
        ProcessingJob.job_type == "recalculate_kpi_values",
        ProcessingJob.subject_id == v.id,
    )).first()
    assert job is not None, "recalculate_kpi_values job was not enqueued on approval"
    from app.workers import _HANDLERS
    _HANDLERS["recalculate_kpi_values"](session, job)
    return v


# ---------- Seed idempotency ----------

def test_kpi_definitions_seed_is_idempotent(session):
    bootstrap(session)
    seed_upload_types(session)
    r1 = seed_kpi_definitions(session)
    r2 = seed_kpi_definitions(session)
    assert r1["kpi_definitions_added"] == 5
    assert r2["kpi_definitions_added"] == 0
    total = session.exec(select(KpiDefinition)).all()
    assert len(total) == 5


# ---------- Real extraction, happy path ----------

def test_energy_extraction_creates_real_kpi_values(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    site = _make_site(session, co, "SITE-A", "Site A")
    uploader = make_user(session, "eu@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "er@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _energy_xlsx([
        ["Site A", "electricity", 4500, "kWh", "Q2 2026"],
        ["Site A", "diesel", 120, "L", "Q2 2026"],
    ])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "energy_data", xlsx)
    _approve(client, session, uploader, reviewer, ds_pid, v_pid)
    v = _run_kpi_job(session, v_pid)

    values = session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v.id)).all()
    assert len(values) == 2

    row1 = next(kv for kv in values if kv.value == 4500)
    assert row1.kpi_code == "energy.consumption"
    assert row1.unit == "kWh"
    assert row1.site_id == site.id
    assert row1.attributes.get("energy_type") == "electricity"
    assert row1.company_id == co.id
    assert row1.source_row_number == 2
    assert row1.reporting_period_start == date(2026, 4, 1)
    assert row1.reporting_period_end == date(2026, 6, 30)

    row2 = next(kv for kv in values if kv.value == 120)
    assert row2.unit == "L"
    assert row2.attributes.get("energy_type") == "diesel"


def test_water_extraction_creates_two_metrics_per_row(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    site = _make_site(session, co, "SITE-B", "Site B")
    uploader = make_user(session, "wu@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "wr@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _water_xlsx([["Site B", "groundwater", 250.5, 80.2, "ML", "Q2 2026"]])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "water_data", xlsx)
    _approve(client, session, uploader, reviewer, ds_pid, v_pid)
    v = _run_kpi_job(session, v_pid)

    values = session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v.id)).all()
    assert len(values) == 2
    codes = {kv.kpi_code: kv.value for kv in values}
    assert codes["water.withdrawal"] == 250.5
    assert codes["water.recycled"] == 80.2
    for kv in values:
        assert kv.site_id == site.id
        assert kv.attributes.get("water_source_type") == "groundwater"


def test_waste_extraction_captures_both_qualifying_attributes(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    uploader = make_user(session, "wau@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "war@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _waste_xlsx([["Site C", "hazardous", 12.4, "landfill", "t", "Q2 2026"]])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "waste_data", xlsx)
    _approve(client, session, uploader, reviewer, ds_pid, v_pid)
    v = _run_kpi_job(session, v_pid)

    values = session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v.id)).all()
    assert len(values) == 1
    kv = values[0]
    assert kv.kpi_code == "waste.generated"
    assert kv.value == 12.4
    assert kv.attributes.get("waste_type") == "hazardous"
    assert kv.attributes.get("disposal_method") == "landfill"
    assert kv.site_id is None


def test_emissions_factor_column_is_never_captured_as_a_value(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    uploader = make_user(session, "emu@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "emr@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _emissions_xlsx([["Site D", "Scope 1", 500, "DEFRA 2024 Grid Factor", "L", "Q2 2026"]])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "emissions_data", xlsx)
    _approve(client, session, uploader, reviewer, ds_pid, v_pid)
    v = _run_kpi_job(session, v_pid)

    values = session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v.id)).all()
    assert len(values) == 1
    kv = values[0]
    assert kv.kpi_code == "emissions.activity_data"
    assert kv.value == 500
    assert kv.attributes.get("emission_scope") == "Scope 1"
    assert "DEFRA" not in str(kv.attributes)
    assert "factor" not in str(kv.attributes).lower()


# ---------- Invalid data / unknown site handling ----------

def test_invalid_numeric_value_is_skipped_not_fatal(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    uploader = make_user(session, "iu@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "ir@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _energy_xlsx([
        ["Site A", "electricity", "not-a-number", "kWh", "Q2 2026"],
        ["Site A", "gas", 200, "kWh", "Q2 2026"],
    ])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "energy_data", xlsx)
    _approve(client, session, uploader, reviewer, ds_pid, v_pid)
    v = _run_kpi_job(session, v_pid)

    values = session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v.id)).all()
    assert len(values) == 1
    assert values[0].value == 200


def test_unmapped_site_text_does_not_fabricate_a_site(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    uploader = make_user(session, "su@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "sr@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _energy_xlsx([["Nonexistent Site", "electricity", 100, "kWh", "Q2 2026"]])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "energy_data", xlsx)
    _approve(client, session, uploader, reviewer, ds_pid, v_pid)
    v = _run_kpi_job(session, v_pid)

    values = session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v.id)).all()
    assert len(values) == 1
    assert values[0].site_id is None


# ---------- Idempotency / duplicate handling ----------

def test_running_extraction_twice_does_not_duplicate(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    uploader = make_user(session, "du@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "dr@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _energy_xlsx([["Site A", "electricity", 300, "kWh", "Q2 2026"]])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "energy_data", xlsx)
    _approve(client, session, uploader, reviewer, ds_pid, v_pid)
    v = _run_kpi_job(session, v_pid)
    first_count = len(session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v.id)).all())
    assert first_count == 1

    job = session.exec(select(ProcessingJob).where(
        ProcessingJob.job_type == "recalculate_kpi_values", ProcessingJob.subject_id == v.id
    )).first()
    from app.workers import _HANDLERS
    _HANDLERS["recalculate_kpi_values"](session, job)

    second_count = len(session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v.id)).all())
    assert second_count == 1


# ---------- Rejected / changes-requested must NOT produce KPI values ----------

def test_rejected_version_produces_no_kpi_values(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    uploader = make_user(session, "reu@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "rer@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _energy_xlsx([["Site A", "electricity", 300, "kWh", "Q2 2026"]])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "energy_data", xlsx)

    r = client.post(f"/api/v1/datasets/{ds_pid}/versions/{v_pid}/reviews",
                    json={"reviewer_user_id": reviewer.id}, headers=auth(uploader))
    rv_pid = r.json()["public_id"]
    r = client.post(f"/api/v1/datasets/{ds_pid}/versions/{v_pid}/reviews/{rv_pid}/decide",
                    json={"decision": "rejected", "note": "Not acceptable."}, headers=auth(reviewer))
    assert r.status_code == 200, r.text

    v = session.exec(select(DatasetVersion).where(DatasetVersion.public_id == v_pid)).first()
    job = session.exec(select(ProcessingJob).where(
        ProcessingJob.job_type == "recalculate_kpi_values", ProcessingJob.subject_id == v.id
    )).first()
    assert job is None
    values = session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v.id)).all()
    assert len(values) == 0


# ---------- New version after changes-requested gets its own separate KPI rows ----------

def test_new_version_after_changes_requested_has_independent_kpi_values(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    uploader = make_user(session, "nvu@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "nvr@d.com", "deloitte", "Reviewer", org=org)

    xlsx_v1 = _energy_xlsx([["Site A", "electricity", 100, "kWh", "Q2 2026"]])
    ds_pid, v1_pid = _submitted_version(client, session, uploader, co, "energy_data", xlsx_v1)
    r = client.post(f"/api/v1/datasets/{ds_pid}/versions/{v1_pid}/reviews",
                    json={"reviewer_user_id": reviewer.id}, headers=auth(uploader))
    rv1_pid = r.json()["public_id"]
    client.post(f"/api/v1/datasets/{ds_pid}/versions/{v1_pid}/reviews/{rv1_pid}/decide",
               json={"decision": "changes_requested", "note": "Fix the numbers."}, headers=auth(reviewer))

    r = client.post(f"/api/v1/datasets/{ds_pid}/versions", headers=auth(uploader))
    assert r.status_code == 201, r.text
    v2_pid = r.json()["public_id"]
    xlsx_v2 = _energy_xlsx([["Site A", "electricity", 999, "kWh", "Q2 2026"]])
    client.post(f"/api/v1/datasets/{ds_pid}/versions/{v2_pid}/files",
               files={"file": ("d2.xlsx", xlsx_v2,
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
               data={"role": "data"}, headers=auth(uploader))
    client.post(f"/api/v1/datasets/{ds_pid}/versions/{v2_pid}/submit", headers=auth(uploader))
    _approve(client, session, uploader, reviewer, ds_pid, v2_pid)
    v2 = _run_kpi_job(session, v2_pid)

    v2_values = session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v2.id)).all()
    assert len(v2_values) == 1
    assert v2_values[0].value == 999

    v1 = session.exec(select(DatasetVersion).where(DatasetVersion.public_id == v1_pid)).first()
    v1_values = session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v1.id)).all()
    assert len(v1_values) == 0


# ---------- API: tenancy and filtering ----------

def test_kpi_values_api_tenant_isolation(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co_a = make_company(session, org, "Company A")
    co_b = make_company(session, org, "Company B")
    uploader_a = make_user(session, "apia@d.com", "deloitte", "Administrator", org=org)
    reviewer_a = make_user(session, "apir@d.com", "deloitte", "Reviewer", org=org)
    client_b = make_user(session, "apib@client.com", "client", "Client Uploader", company=co_b)

    xlsx = _energy_xlsx([["Site A", "electricity", 100, "kWh", "Q2 2026"]])
    ds_pid, v_pid = _submitted_version(client, session, uploader_a, co_a, "energy_data", xlsx)
    _approve(client, session, uploader_a, reviewer_a, ds_pid, v_pid)
    _run_kpi_job(session, v_pid)

    r = client.get("/api/v1/kpi-values", headers=auth(client_b))
    assert r.status_code == 200
    assert r.json()["total"] == 0

    r = client.get("/api/v1/kpi-values", headers=auth(uploader_a))
    assert r.status_code == 200
    assert r.json()["total"] == 1
    item = r.json()["items"][0]
    assert "public_id" in item and item["public_id"].startswith("kv_")
    assert item["dataset_public_id"] == ds_pid


def test_kpi_values_api_filters_by_kpi_code(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    uploader = make_user(session, "fu@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "fr@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _water_xlsx([["Site A", "groundwater", 50, 20, "ML", "Q2 2026"]])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "water_data", xlsx)
    _approve(client, session, uploader, reviewer, ds_pid, v_pid)
    _run_kpi_job(session, v_pid)

    r = client.get("/api/v1/kpi-values?kpi_code=water.recycled", headers=auth(uploader))
    assert r.json()["total"] == 1
    assert r.json()["items"][0]["kpi_code"] == "water.recycled"
    assert r.json()["items"][0]["value"] == 20


def test_kpi_definitions_api_no_formula_fields_exposed(client, session):
    org = _seed(session)
    admin = make_user(session, "kdu@d.com", "deloitte", "Administrator", org=org)
    r = client.get("/api/v1/kpi-definitions?upload_type_code=energy_data", headers=auth(admin))
    assert r.status_code == 200
    assert len(r.json()) == 1
    item = r.json()[0]
    assert item["code"] == "energy.consumption"
    assert set(item.keys()) == {"code", "display_name", "unit_hint", "upload_type_code", "data_type", "version", "is_active"}


# ---------- HARDENING PASS: Item 1 — no internal ids exposed ----------

def test_kpi_values_api_never_exposes_raw_internal_ids(client, session, tmp_path):
    """Explicit, exhaustive check: not a single raw internal integer id
    (dataset_id, dataset_version_id, source_file_id, site_id) appears
    anywhere in the API response — only public_ids and the one
    deliberate exception (company_id, which has no public_id anywhere
    in this codebase)."""
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    site = _make_site(session, co, "SITE-X", "Site X")
    uploader = make_user(session, "hidu@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "hidr@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _energy_xlsx([["Site X", "electricity", 100, "kWh", "Q2 2026"]])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "energy_data", xlsx)
    _approve(client, session, uploader, reviewer, ds_pid, v_pid)
    v = _run_kpi_job(session, v_pid)
    kv = session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v.id)).first()

    r = client.get("/api/v1/kpi-values", headers=auth(uploader))
    item = r.json()["items"][0]

    # The exact set of keys must match — no extra raw-id field slipped in.
    assert set(item.keys()) == {
        "public_id", "dataset_public_id", "dataset_version_public_id",
        "source_file_public_id", "source_row_number", "company_id",
        "site_public_id", "kpi_code", "kpi_definition_version", "value",
        "unit", "attributes", "reporting_period_start", "reporting_period_end",
        "created_at",
    }
    # Every id-shaped field must be a public_id string, not a raw int,
    # except company_id (documented exception -- no public_id exists).
    assert item["public_id"].startswith("kv_")
    assert item["dataset_public_id"].startswith("ds_")
    assert item["dataset_version_public_id"].startswith("dv_")
    assert item["source_file_public_id"].startswith("df_")
    assert item["site_public_id"] == site.public_id
    assert item["site_public_id"].startswith("st_")
    assert isinstance(item["company_id"], int)  # the one documented exception


def test_kpi_values_api_site_filter_uses_public_id_not_raw_int(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    site = _make_site(session, co, "SITE-Y", "Site Y")
    uploader = make_user(session, "sfu@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "sfr@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _energy_xlsx([["Site Y", "gas", 77, "kWh", "Q2 2026"]])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "energy_data", xlsx)
    _approve(client, session, uploader, reviewer, ds_pid, v_pid)
    _run_kpi_job(session, v_pid)

    r = client.get(f"/api/v1/kpi-values?site_public_id={site.public_id}", headers=auth(uploader))
    assert r.json()["total"] == 1
    assert r.json()["items"][0]["value"] == 77


# ---------- HARDENING PASS: Item 2 — kpi_definition_version provenance ----------

def test_kpi_value_records_the_active_definition_version(client, session, tmp_path):
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    uploader = make_user(session, "vpu@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "vpr@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _energy_xlsx([["Site A", "electricity", 100, "kWh", "Q2 2026"]])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "energy_data", xlsx)
    _approve(client, session, uploader, reviewer, ds_pid, v_pid)
    v = _run_kpi_job(session, v_pid)

    kv = session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v.id)).first()
    definition = session.exec(select(KpiDefinition).where(KpiDefinition.code == "energy.consumption")).first()
    # Not just "equals 1" (which would pass even if the stamping logic were
    # a no-op default) -- genuinely equals whatever the real, current
    # active definition's version is, proving it was actually resolved.
    assert kv.kpi_definition_version == definition.version

    r = client.get("/api/v1/kpi-values", headers=auth(uploader))
    assert r.json()["items"][0]["kpi_definition_version"] == definition.version


def test_kpi_value_version_unambiguously_identifies_the_definition_even_if_a_new_version_is_added_later(session):
    """If a second version of the same code is added later (e.g. a future
    unit_hint change), an EXISTING KpiValue's (code, version) pair must
    still point at the ORIGINAL definition it was extracted under, not
    silently appear to belong to the new one."""
    org = bootstrap(session)
    seed_upload_types(session)
    seed_kpi_definitions(session)
    d_v1 = session.exec(select(KpiDefinition).where(KpiDefinition.code == "energy.consumption")).first()
    assert d_v1.version == 1

    # Simulate a future definition change: version 2 of the same code.
    ut = session.exec(select(UploadType).where(UploadType.code == "energy_data")).first()
    d_v2 = KpiDefinition(code="energy.consumption", display_name="Energy Consumption (revised)",
                         unit_hint="MWh", upload_type_id=ut.id, version=2)
    session.add(d_v2); session.commit(); session.refresh(d_v2)

    # An existing KpiValue stamped with version=1 must be resolvable back
    # to the ORIGINAL (kWh) definition, not the new (MWh) one, via the
    # (code, version) compound identity — proving the two are genuinely
    # distinguishable rather than colliding on code alone.
    old_def = session.exec(select(KpiDefinition).where(
        KpiDefinition.code == "energy.consumption", KpiDefinition.version == 1
    )).first()
    new_def = session.exec(select(KpiDefinition).where(
        KpiDefinition.code == "energy.consumption", KpiDefinition.version == 2
    )).first()
    assert old_def.unit_hint == "kWh"
    assert new_def.unit_hint == "MWh"
    assert old_def.id != new_def.id


# ---------- HARDENING PASS: Item 3 — never zero-fill, explicit blank-cell test ----------

def test_blank_cell_is_skipped_never_interpreted_as_zero(client, session, tmp_path):
    """Distinct from the existing 'not-a-number' string test -- this
    covers a genuinely BLANK cell (None), the most common real-world
    case of missing data, and proves it is skipped rather than silently
    becoming a fabricated 0."""
    _set_storage(tmp_path)
    org = _seed(session)
    co = make_company(session, org, "Co")
    uploader = make_user(session, "bcu@d.com", "deloitte", "Administrator", org=org)
    reviewer = make_user(session, "bcr@d.com", "deloitte", "Reviewer", org=org)

    xlsx = _energy_xlsx([
        ["Site A", "electricity", None, "kWh", "Q2 2026"],  # blank value cell
        ["Site A", "gas", 50, "kWh", "Q2 2026"],
    ])
    ds_pid, v_pid = _submitted_version(client, session, uploader, co, "energy_data", xlsx)
    _approve(client, session, uploader, reviewer, ds_pid, v_pid)
    v = _run_kpi_job(session, v_pid)

    values = session.exec(select(KpiValue).where(KpiValue.dataset_version_id == v.id)).all()
    assert len(values) == 1  # NOT 2 -- the blank row must not become a 0.0 row
    assert values[0].value == 50
    # Explicitly confirm no row with value 0.0 was ever created for the
    # blank cell -- this is the exact failure mode being guarded against.
    assert not any(kv.value == 0.0 for kv in values)


# ---------- HARDENING PASS: Item 4 — genuine DB-level constraint safety ----------

def test_database_level_unique_violation_is_caught_and_rolled_back_cleanly(session, tmp_path):
    """Distinct from test_running_extraction_twice_does_not_duplicate,
    which only exercises the application-level pre-check (the early
    'any rows already exist for this version' short-circuit). This test
    bypasses that pre-check entirely and forces a genuine database-level
    unique-constraint violation, proving the DEEPER safety net (the
    try/except around session.commit()) actually works -- this is the
    real protection against a genuine concurrent-extraction race, not
    just the single-threaded convenience check."""
    from app.models.kpi import KpiValue as KV

    _set_storage(tmp_path)
    org = bootstrap(session)
    seed_upload_types(session)
    seed_kpi_definitions(session)
    co = make_company(session, org, "Co")
    uploader = make_user(session, "raceu@d.com", "deloitte", "Administrator", org=org)

    # Build a dataset/version/file directly (not through the API) so we
    # can manufacture the exact conflicting row before calling extraction,
    # bypassing the pre-check without needing real thread concurrency.
    ut = session.exec(select(UploadType).where(UploadType.code == "energy_data")).first()
    ds = Dataset(company_id=co.id, upload_type_id=ut.id,
                reporting_period_start=date(2026, 4, 1), reporting_period_end=date(2026, 6, 30),
                created_by=uploader.id)
    session.add(ds); session.commit(); session.refresh(ds)
    v = DatasetVersion(dataset_id=ds.id, version_number=1, status="approved", uploaded_by=uploader.id)
    session.add(v); session.commit(); session.refresh(v)

    from app.storage import get_storage
    from app.storage.factory import build_storage_key
    xlsx = _energy_xlsx([["Site A", "electricity", 100, "kWh", "Q2 2026"]])
    key = build_storage_key(co.id, ds.public_id, 1, "df_race", "d.xlsx")
    stored = get_storage().put(key, io.BytesIO(xlsx),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    df = DatasetFile(public_id="df_race_test", dataset_version_id=v.id, role="data",
                     storage_key=key, original_filename="d.xlsx", mime_type=stored.mime_type,
                     size_bytes=stored.size_bytes, sha256_checksum=stored.sha256_checksum,
                     uploaded_by=uploader.id)
    session.add(df); session.commit(); session.refresh(df)

    # Manufacture the EXACT conflicting row a real extraction run would
    # produce (same dataset_version_id/source_file_id/source_row_number/
    # kpi_code) -- simulating that a concurrent run already committed it
    # a moment before ours.
    session.add(KV(
        dataset_id=ds.id, dataset_version_id=v.id, source_file_id=df.id,
        source_row_number=2, company_id=co.id, site_id=None,
        kpi_code="energy.consumption", kpi_definition_version=1,
        value=999.0, unit="kWh", attributes={},
        reporting_period_start=ds.reporting_period_start,
        reporting_period_end=ds.reporting_period_end,
    ))
    session.commit()

    # Call the lower-level write function DIRECTLY -- this genuinely
    # bypasses the higher-level "already extracted" pre-check (which
    # would otherwise short-circuit before ever reaching the commit),
    # so this actually exercises the DB-level constraint + rollback path,
    # not just the earlier, already-separately-tested application guard.
    from app.services.kpi_extraction_service import _write_kpi_values
    conflicting_insert = [KV(
        dataset_id=ds.id, dataset_version_id=v.id, source_file_id=df.id,
        source_row_number=2, company_id=co.id, site_id=None,
        kpi_code="energy.consumption", kpi_definition_version=1,
        value=100.0, unit="kWh", attributes={"energy_type": "electricity"},
        reporting_period_start=ds.reporting_period_start,
        reporting_period_end=ds.reporting_period_end,
    )]
    result = _write_kpi_values(session, conflicting_insert, rows_skipped=0)

    # Must report the clean, understood outcome -- not raise, not corrupt.
    assert result["reason"] == "already_extracted_concurrently"
    assert result["rows_written"] == 0

    # The session must be left completely usable afterward -- proving no
    # dirty/broken transaction state leaked out, matching this module's
    # documented safety design. A follow-up query must succeed cleanly.
    remaining = session.exec(select(KV).where(KV.dataset_version_id == v.id)).all()
    assert len(remaining) == 1  # still exactly the ORIGINAL manufactured row
    assert remaining[0].value == 999.0  # the conflicting insert never landed

    # And the session must still be usable for further real work --
    # e.g. a genuinely new, non-conflicting row must insert fine.
    session.add(KV(
        dataset_id=ds.id, dataset_version_id=v.id, source_file_id=df.id,
        source_row_number=3, company_id=co.id, site_id=None,
        kpi_code="energy.consumption", kpi_definition_version=1,
        value=42.0, unit="kWh", attributes={},
        reporting_period_start=ds.reporting_period_start,
        reporting_period_end=ds.reporting_period_end,
    ))
    session.commit()  # must not raise
    final_count = len(session.exec(select(KV).where(KV.dataset_version_id == v.id)).all())
    assert final_count == 2
